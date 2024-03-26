import os

from fastapi import APIRouter, BackgroundTasks, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook

router = APIRouter()


# Define a model to receive the file path
class FileInput(BaseModel):
    file_path: str
    vendor_id: str
    store_id: str
    import_type: str = None
    tenant_name: str = None


def split_excel_chunks(file_path: str, chunk_size: int = 1000):
    workbook = load_workbook(file_path, read_only=True)

    # Get the active worksheet
    sheet = workbook.active

    # Create an ExcelWriter object for the first chunk
    output_file_template = 'split_files/output_chunk_{}.xlsx'
    output_file_number = 1
    output_file_path = output_file_template.format(output_file_number)
    output_workbook = Workbook()
    output_sheet = output_workbook.active

    # Counter for tracking rows in the current chunk
    current_row_count = 0

    # Iterate through rows
    for row in sheet.iter_rows(values_only=True):

        # Write the row to the output sheet
        output_sheet.append(row)
        current_row_count += 1

        # Check if we have reached the chunk size
        if current_row_count >= chunk_size:
            # Save the current chunk to a new file
            output_workbook.save(output_file_path)

            # Close the current workbook and sheet
            output_workbook.close()

            # Increment the output file number for the next chunk
            output_file_number += 1
            output_file_path = output_file_template.format(output_file_number)

            # Create a new workbook and sheet for the next chunk
            output_workbook = Workbook()
            output_sheet = output_workbook.active

            # Reset the row count for the new chunk
            current_row_count = 0

    # Save the last chunk to a file
    output_workbook.save(output_file_path)

    # Close the last workbook
    output_workbook.close()

    # Close the original workbook
    workbook.close()

    print("--------------------Splitting complete--------------------")
    return 1


# API endpoint to receive file path and start the conversion
@router.post("/split-file")
async def split_excel_to_chunks(file: FileInput, background_tasks: BackgroundTasks):
    # check file.file_path exists or not
    if not os.path.isfile(file.file_path):
        raise HTTPException(status_code=400, detail="File does not exist!")

    background_tasks.add_task(split_excel_chunks, file.file_path, 1000)
    return {"message": "output_file_path"}


# API endpoint to get total record count without waiting for split
@router.post("/file-count")
async def get_total_count(file_input: FileInput):
    workbook = load_workbook(file_input.file_path, read_only=True)

    # Get the active worksheet
    sheet = workbook.active

    # Initialize a count variable
    total_count = 0

    # Iterate through rows and count
    for _ in sheet.iter_rows(values_only=True):
        total_count += 1
        print(total_count)

    # Close the workbook when done
    workbook.close()
    return {"total_count": total_count}
